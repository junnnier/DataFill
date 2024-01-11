from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.drawing.image import Image as openpyxl_image
import os
import copy
import yaml


def load_template(path, read_only=False):
    workbook = load_workbook(path, data_only=read_only)
    return workbook, workbook.get_sheet_names()


def write_data(workbook, sheet_name, define_data, input_data):
    position = define_data["position"]
    # 图片/文本数据
    if isinstance(input_data, openpyxl_image):
        workbook[sheet_name].add_image(copy.deepcopy(input_data), position)
    else:
        if define_data["replace"] is None:
            workbook[sheet_name][position] = input_data
        else:
            workbook[sheet_name][position] = str(workbook[sheet_name][position].value).replace(define_data["replace"], str(input_data))
        if "font" in define_data.keys():
            workbook[sheet_name][position].font = Font(**define_data["font"])
        if "number_format" in define_data.keys():
            workbook[sheet_name][position].number_format = define_data["number_format"]


def data_process(workbook, sheet_name_list, data, config):
    rule = config["rule"]
    # 对每个数值操作
    for index, item in enumerate(data):
        if index in rule.keys():
            # 根据设定的规则填写数据
            define_data_list = rule[index]
            for define_data in define_data_list:
                write_data(workbook, sheet_name_list[define_data["sheet"]], define_data, item)
    return workbook


def post_process(workbook, sheet_name_list, data):
    # 对每个数据进行处理
    for key, values_list in data.items():
        # 根据设定的规则填写数据
        for values in values_list:
            write_data(workbook, sheet_name_list[values[0]], values, key)
    return workbook


def check_image(worksheet):
    result = {}
    # 获取图片所在的行和列
    for image in worksheet._images:
        row = image.anchor._from.row
        column = image.anchor._from.col
        result[row] = [column, image]
    return result


def execution(data_path, template_path, config_path, save_dir):
    # 读取配置
    with open(config_path, "r", encoding="utf-8") as f:
        config = yaml.safe_load(f)
    print("load config file success.")
    # 读取数据
    data_workbook, data_sheet_list = load_template(data_path, read_only=True)
    print("load data file success.")
    data_sheet = data_workbook.active
    # 检查是否存在image元素
    sheet_image_info = check_image(data_sheet)
    # 对每行数据操作
    data_number = data_sheet.max_row - 1
    for row_num, row_value in enumerate(data_sheet.values):
        row_value = list(row_value)
        # 跳过表头数据
        if row_num == 0:
            continue
        try:
            # 如果该行有图片数据，提取到对应列
            if sheet_image_info and row_num in sheet_image_info.keys():
                image_info = sheet_image_info[row_num]
                row_value[image_info[0]] = image_info[1]
            # 加载模板文件
            template_workbook, template_sheet_list = load_template(template_path)
            # 填写数据
            new_workbook = data_process(template_workbook, template_sheet_list, row_value, config)
            # 是否进行后处理
            if config["rule_post"]:
                new_workbook = post_process(new_workbook, template_sheet_list, config["rule_post"])
            # 保存路径
            save_path = os.path.join(save_dir, "{}.xlsx".format(row_value[0]))
            # 保存
            new_workbook.save(save_path)
            template_workbook.close()
            print("[{}/{}] success, the file save to: {}".format(row_num, data_number, save_path))
        except Exception as e:
            print("[{}/{}] fail, error: {}".format(row_num, data_number, e))
    # 关闭
    data_workbook.close()

